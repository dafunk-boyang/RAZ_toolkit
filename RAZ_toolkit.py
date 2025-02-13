import datetime
import os.path
import selenium.common.exceptions
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
import pandas as pd
from bs4 import BeautifulSoup
from scipy.stats import chi2_contingency
import matplotlib.pyplot as plt
import numpy as np
from fpdf import FPDF
import openpyxl

class Driver:

    def __init__(self):
        """
        Sets up a driver
        """
        self.run_number = 0

        # Create a Chrome instance and set window size
        self.driver = webdriver.Chrome()
        self.driver.set_window_size(1294, 696)

    def login(self, username, password):
        self.username = username
        self.password = password
        # find login input fields and input credentials
        try:
            self.driver.get("https://accounts.learninga-z.com/ng/member/login?siteAbbr=kaz")
            WebDriverWait(self.driver, 10).until(lambda x: x.find_element(By.NAME, "username"))
        except selenium.common.exceptions.TimeoutException:
            print("Page Timed out, please try again")

        self.driver.find_element(By.NAME, "username").send_keys(username)
        self.driver.find_element(By.CSS_SELECTOR, ".password-input > .ng-untouched").send_keys(password)
        self.driver.find_element(By.CSS_SELECTOR, ".accounts-primary-btn > .mat-button-wrapper").click()
        if self.driver.find_elements(By.XPATH, '//*[@id="alert-description"]/div[1]/i'):
            print("Incorrect Username or Password")
        else:
            try:
                WebDriverWait(self.driver, 10).until(
                    lambda x: x.find_element(By.XPATH, '//*[@id="manageStudents"]/laz-popover-target/button'))
                WebDriverWait(self.driver, 10).until(lambda x: x.find_element(By.ID, "onetrust-accept-btn-handler"))
                if self.driver.find_element(By.XPATH, '//*[@id="onetrust-accept-btn-handler"]'):
                    self.driver.find_element(By.XPATH, '//*[@id="onetrust-accept-btn-handler"]').click()
            except selenium.common.exceptions.NoSuchElementException:
                print("Could not find checkpoint element")
        print("Web driver initialized")

    def get_student_list(self):
        self.driver.get("https://kidsa-z.com/main/Dashboard#!/overview")
        try:
            WebDriverWait(self.driver, 10).until(
                lambda x: x.find_element(By.XPATH, '//*[@id="manageStudents"]/laz-popover-target/button'))
        except selenium.common.exceptions.TimeoutException:
            print("Timed out, please try again")
        except selenium.common.exceptions.NoSuchElementException:
            print("Could not find checkpoint element")

        self.driver.find_element(By.XPATH, '//*[@id="manageStudents"]/laz-popover-target/button').click()
        page_source = self.driver.page_source

        # Parse the HTML using BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        # Find all 'a' tags with class 'dropdownNav_pageLabel'
        name_elements = soup.find_all('a', class_='dropdownNav_pageLabel')
        student_list = [element.get_text(strip=True).replace('Profile', '') for element in name_elements]
        students = pd.DataFrame(student_list, columns=['Name'])

        print("Student list acquired")
        return students

    def get_student_accuracy_data(self, names):
        self.driver.get('https://www.kidsa-z.com/main/classreports#!/class/skill/report')
        sums = []
        self.driver.find_element(By.ID, "dateDropDown").click()
        self.driver.find_element(By.XPATH, '//*[@id="mat-option-43"]/span').click()


        for i, row in names.iterrows():
            try:
                self.driver.find_element(By.XPATH, '//*[@id="mat-select-8"]/div/div[2]').click()
                self.driver.find_element(By.XPATH, f'//*[@id="mat-option-{15+i}"]/span/span').click()
                WebDriverWait(self.driver, 3).until(
                    lambda x: x.find_element(By.XPATH, '//class-skill-report/div/div/table'))
            except selenium.common.exceptions.TimeoutException:
                print(f"Table not found for {row['Name']}.")
                sums.append(None)
                continue

            page_source = self.driver.page_source
            # Parse the HTML using BeautifulSoup
            soup = BeautifulSoup(page_source, 'html.parser')
            # Find the div with the specified class
            table = soup.find('table', {'class': 'table-data table-full table-reportsTable table-orion'})
            if table:
                # Extract column headers
                headers = [th.text.strip() for th in table.find_all('th')]
                # Extract rows
                data = []
                for row in table.find_all('tr')[1:]:  # Skip the header row
                    row_data = [td.text.strip() for td in row.find_all('td')]
                    data.append(row_data)

            # Create a Pandas DataFrame
            accuracy_data = pd.DataFrame(data, columns=headers)
            sums.append(accuracy_data['Correct'].astype(int).sum() / accuracy_data['Total'].astype(int).sum())
            accuracy_data['Skill'] = accuracy_data['Skill'].apply(
                lambda x: ' '.join(x.split('\n\n')[:-1]) if '\n\n' in x else x)

        names['Overall Skills Accuracy'] = sums
        accuracies = names
        accuracies.rename(columns={'Overall Skills Accuracy': f'{datetime.date.today()}'}, inplace=True)
        return accuracies

    def navigate_to_student_activity_page(self, student_num):
        self.driver.get("https://www.kidsa-z.com/main/classreports#!/class/activity?subject=")
        if self.run_number > 0:
            self.driver.find_element(By.ID, "studentsDropDown").click()
            self.driver.find_element(By.ID, f'{student_num}-student-option').click()
            WebDriverWait(self.driver, 10).until(lambda x: x.find_element(By.ID, "studentData"))
            if self.driver.find_element(By.XPATH, "//div/ul/li[3]/button"):
                try:
                    self.driver.find_element(By.XPATH, "//div/ul/li[3]/button").click()
                    WebDriverWait(self.driver, 5).until(lambda x: x.find_element(By.ID, "studentData"))
                except selenium.common.exceptions.NoSuchElementException:
                    print(f"Could not find data element for student {student_num}")
            else:
                print(f"Student number {student_num} didn't have a lot of data")
                self.run_number += 1
        else:
            try:
                WebDriverWait(self.driver, 30).until(lambda x: x.find_element(By.ID, "dateDropDown"))
                self.driver.find_element(By.ID, "dateDropDown").click()
                self.driver.find_element(By.XPATH, '//*[@id="mat-option-4"]/span').click()
                self.driver.find_element(By.ID, "studentsDropDown").click()
                self.driver.find_element(By.ID, f'{student_num}-student-option').click()
                WebDriverWait(self.driver, 30).until(lambda x: x.find_element(By.ID, "studentData"))
                self.driver.find_element(By.XPATH, "//div/ul/li[3]/button").click()
                WebDriverWait(self.driver, 30).until(lambda x: x.find_element(By.ID, "studentData"))
                self.run_number += 1
            except selenium.common.exceptions.NoSuchElementException:
                print(f"Could not find checkpoint element in student {student_num}")

    def get_cycle_data(self):
        """
        Scrapes data from given page
        :param self: browser driver from the class initialization
        :return df: dataframe scraped from site
        """
        # read page source from selenium
        page_source = self.driver.page_source

        # Parse the HTML using BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        # Find the table with the ID "studentData"
        table = soup.find('table', {'id': 'studentData'})

        # Extract column headers
        headers = [th.text.strip() for th in table.find_all('th')]

        # Extract rows
        data = []
        for row in table.find_all('tr')[1:]:  # Skip the header row
            row_data = [td.text.strip() for td in row.find_all('td')]
            data.append(row_data)

        # Create a Pandas DataFrame
        df = pd.DataFrame(data, columns=headers)
        return df

    def get_weekly_readers(self):
        self.driver.get('https://www.kidsa-z.com/main/classreports#!/class/activity?subject=')

        try:
            self.driver.find_element(By.ID, 'dateDropDown').click()
            self.driver.find_element(By.XPATH, '//*[@id="mat-option-29"]/span').click()
            WebDriverWait(self.driver, 3).until(
                    lambda x: x.find_element(By.XPATH, '//*[@id="classActivityTable"]'))
        except selenium.common.exceptions.TimeoutException:
            print(f"Table not found.")

        page_source = self.driver.page_source
        # Parse the HTML using BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')
        # Find the div with the specified class
        table = soup.find('table', {'class': 'table-data'})
        if table:
            # Extract column headers
            headers = [th.text.strip() for th in table.find_all('th')]
            # Extract rows
            data = []
            for row in table.find_all('tr')[1:]:  # Skip the header row
                row_data = [td.text.strip() for td in row.find_all('td')]
                data.append(row_data)
        else:
            print("Table not found.")

        activity_data = pd.DataFrame(data, columns=headers)

        return activity_data

    def quit(self):
        """
        Quits web driver
        """
        self.driver.quit()
        pass


class DataHandler:

    def reading_cycle_cleaner(self, dirty_data):
        """
        Cleans dataframe of superfluous instances to our analysis.
        :param dirty_data: Uncleaned data from the site.
        :return clean_data: Data needed for analysis, clean of newline tags and other junk.
        """
        cycle_targets = ['Listen', 'Read', 'Quiz', 'Practice Recording']
        cycle_instances = dirty_data[dirty_data['Type'].isin(
            cycle_targets)]  # & (                               # cleaned up superfluous statements
        # (dirty_data['Location'] == 'Reading Room') | (dirty_data['Location'] == 'Level Up!'))]
        clean_data = cycle_instances.copy()
        clean_data['Title'] = clean_data['Title'].str.split('\n').str[0]
        # clean_data = clean_data[clean_data.duplicated(subset='Title', keep=False)].copy()
        clean_data.drop(columns=['Info', 'Stars'], inplace=True)

        # mask for decodables
        mask = clean_data['Title'].isin(title_all)
        # Filter the DataFrame using the mask to get rows where 'Title' is in title_all
        filtered_data = clean_data[mask]
        # drop masked data
        clean_data = clean_data.drop(clean_data[mask].index)

        return clean_data, filtered_data

    def summarize_titles(self, dataframe):
        """
        Summarizes the interaction data from the RAZ site into key interactions per book title.
        :param dataframe: DataFrame from the site (hopefully has been passed through the cleaner).
        :return summary_df: Summary data per each book title.
        """

        # Select relevant columns
        selected_columns = ['Title', 'Type', 'Date']  # Add 'Date' column
        selected_data = dataframe[selected_columns]

        # Initialize a dictionary to store results
        title_summary = {}

        # Iterate through the data and count occurrences of each type for each title
        for _, row in selected_data.iterrows():
            title = row['Title']
            current_type = row['Type']
            current_date = row['Date']  # New: get the current date

            # If the title is not in the summary dictionary, add it
            if title not in title_summary:
                title_summary[title] = {
                    'Listen': 0,
                    'Read': 0,
                    'Quiz': 0,
                    'Practice Recording': 0,
                    'Date': '2020-01-01'  # Initialize date with a default value
                }

            # Increment the count for the current type
            title_summary[title][current_type] += 1

            # Update the date if it's more recent
            title_summary[title]['Date'] = current_date

        # Create a DataFrame from the summary dictionary
        summary_df = pd.DataFrame.from_dict(title_summary, orient='index')
        summary_df.sort_values(by='Date', ascending=False, inplace=True)
        summary_df['Date'] = pd.to_datetime(summary_df['Date']).dt.date

        return summary_df


# From 502 Project
# known in file structure as visualizer.py
class Plots:

    def calculate_chi_square_for_dataframe(self, df):
        """
        Calculate Chi-square statistics for each row in a Dataframe.
        :param df: Dataframe containing observed frequencies for each category.
        :return results: DataFrame containing Chi-square statistics and p-values for each row.
        """
        expected_values = [1, 1, 3, 1]
        result_list = []  # Use a list to store results

        for idx, row in df[['Listen', 'Read', 'Quiz', 'Practice Recording']].iterrows():
            observed_values = row.tolist()
            chi2, p_value, _, _ = chi2_contingency([observed_values, expected_values], correction=False)
            result_list.append({'Title': idx, 'Chi-square': chi2, 'P-value': p_value})

        results = pd.DataFrame(result_list)  # Convert the list to a DataFrame
        return results

    def plot_p_values(self, df, summary_df, name):
        """
        Plot P-values to show adherence to a 'model' RAZ reading cycle.
        :param df: DataFrame with Chi-square and P-value results.
        :param summary_df: DataFrame with summary data.
        :return None: (plots the graph).
        """

        # Line graph long term history
        line = plt.plot(summary_df['Date'].tolist()[::-1], df['P-value'].tolist()[::-1])
        plt.title(f"{name}'s RAZ Reading Cycles")
        plt.xticks(summary_df['Date'].tolist()[::-10])
        plt.ylabel('Reading Cycle Quality')
        plt.yticks([0.2, 0.5, 0.8], ['Needs Improvement', 'Not Bad', 'Great RAZ Cycle!'])
        plt.savefig('line.png')
        plt.clf()

        # Bar graph
        bar = plt.bar(x=np.arange(len(df['Title'].iloc[0:5])), height='P-value', data=df[0:5])
        plt.title(f"{name}'s RAZ Reading Cycles")
        plt.ylabel('Reading Cycle Quality')
        plt.xticks(np.arange(5), summary_df.index[0:5], ha='right', rotation=45, fontsize=10)
        plt.yticks([0.2, 0.5, 0.8], ['Needs Improvement', 'Not Bad', 'Great RAZ Cycle!'])
        plt.savefig('bar.png')
        plt.clf()
        pass

    def plot_tab(self, summary_df):


        # Table
        table_data = summary_df.transpose().iloc[:, :5]  # Transpose the summary_df for table
        # print(table_data.to_string())
        tab = plt.table(cellText=table_data.values[:-1],  # Exclude the last row (which contains dates)
                        rowLabels=table_data.index[:-1],
                        cellLoc='center',
                        loc='right',
                        bbox=[0, -1, 1, 0.5],
                        colLabels=None)
        tab.auto_set_font_size(False)
        plt.subplots_adjust(left=0.274, bottom=0.488, right=0.988, top=0.942)
        plt.tight_layout()
        plt.savefig('tab.png')
        plt.clf()
        pass

class FPDF(FPDF):

    def header(self):
        self.set_font("Arial", "U", 16)
        self.cell(0, 5, "Mr. Everett's RAZ toolkit student report", align="C", border=0)
        # self.set_xy(10, 30)
        return

    def footer(self):
        self.set_y(-15)
        pageNum = self.page_no()
        self.cell(0, 10, str(pageNum), align="R")
        return

    def add_image(self, image_path, x, y, w, h):
        self.image(image_path, x, y, w, h)
        # Move the cursor to the correct position after the image
        self.set_y(y + h + 10)  # Adjust 10 as needed for spacing

    def create_table(self, dataframe):
        self.set_font('Arial', 'B', 12)
        col_width = self.w / (len(dataframe.columns) + 2)  # +2 to account for the index column
        row_height = self.font_size * 1.5

        # Table header with index column
        self.cell(col_width, row_height, 'Index', border=1)
        for column in dataframe.columns:
            self.cell(col_width, row_height, column, border=1)
        self.ln(row_height)

        # Table rows
        self.set_font('Arial', '', 12)
        for idx, row in dataframe.iterrows():
            max_height = 0
            cell_data = [str(idx)] + [str(item) for item in row]
            x_start = self.get_x()
            y_start = self.get_y()
            cell_height = []

            # Calculate the height needed for each cell and find the max height
            for item in cell_data:
                self.multi_cell(col_width, row_height, item, border=0)
                cell_height.append(self.get_y() - y_start)
                max_height = max(max_height, self.get_y() - y_start)
                self.set_xy(x_start + col_width, y_start)
                x_start += col_width

            # Print each cell again with the calculated max height
            x_start = self.get_x() - col_width * len(cell_data)
            for i, item in enumerate(cell_data):
                self.multi_cell(col_width, row_height, item, border=1)
                self.set_xy(x_start + col_width * (i + 1), y_start)

            self.ln(max_height)


def student_data(user='Everett302', pword='gtyhps302'):
    d = Driver()
    d.login(user, pword)
    names = d.get_student_list()
    counter = 0
    slackers = []
    pdf = FPDF("P", "mm", "A4")

    def process_student(num):
        d.navigate_to_student_activity_page(num)
        cycle_data = d.get_cycle_data()

        data_thingy = DataHandler()
        clean, decodable = data_thingy.reading_cycle_cleaner(cycle_data)
        summaries = data_thingy.summarize_titles(clean)
        visualize_data = Plots()
        results_df = visualize_data.calculate_chi_square_for_dataframe(summaries)
        visualize_data.plot_p_values(results_df, summaries, names.iloc[num]['Name'])

        pdf.add_page()
        pdf.set_font("Arial", "", 12)
        pdf.set_xy(10, 10)
        pdf.cell(len(names.iloc[num]['Name']) * 2.5, 6, names.iloc[num]['Name'], 1)
        # pdf.image('line.png', w=150, x=30, y=20)
        pdf.add_image('bar.png', w=150, x=30, y=10, h=150)
        pdf.create_table(summaries[0:5])
        if len(decodable) > 0:
            decodable_summaries = data_thingy.summarize_titles(decodable)
            pdf.create_table(decodable_summaries[0:5])
            # visualize_data.plot_tab(decodable_summaries)
            # pdf.image('tab.png', w=150, x=30, y=200)

    os.system('osascript -e \'tell application "Terminal" to activate\'')
    print(names)
    student_list = []
    user_input = input("Input student number to select to create report\n")
    student_list.append(int(user_input))
    while user_input.upper() != 'Q':
        print(names)
        user_input = input("Input another student number to select to create another report, or q to quit and generate reports\n")
        if user_input.upper() == 'Q':
            break
        else:
            try:
                student_list.append(int(user_input))
            except ValueError:
                print("Invalid Student Number\n")

    for s in student_list:
        process_student(s)


    for slacker in slackers:
        print(f'Data for {slacker} not found')

    print(f'{counter} student reports created')
    pdf.output(f'{datetime.date.today()} {user} Report.pdf')
    print("Student Data report complete")


def class_data(user='Everett302', pword='gtyhps302'):
    d = Driver()
    d.login(user, pword)
    j = d.get_student_accuracy_data(d.get_student_list())
    print(j)

    current_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(current_dir)
    file_path = os.path.join(parent_dir, f'{datetime.date.today()}_{user}_Accuracy_Report.csv')
    j.to_csv(file_path, index=False)

    print(f'{datetime.date.today()}_{user}_Accuracy_Report.csv Created')
    d.quit()


def activity_sorting(data):
    data['Login Time'] = data['Login Time'].apply(time_delta_to_minutes)
    data = data.sort_values('Login Time', ascending=True)
    return data


def time_delta_to_minutes(delta_str):
    parts = delta_str.split()
    total_minutes = 0
    for part in parts:
        if part.endswith('d'):
            total_minutes += int(part[:-1]) * 24 * 60  # Convert days to minutes
        elif part.endswith('h'):
            total_minutes += int(part[:-1]) * 60  # Convert hours to minutes
        elif part.endswith('m'):
            total_minutes += int(part[:-1])  # Add minutes
    time = pd.to_timedelta(total_minutes, unit='m')
    return time


def weekly_readers(user='Everett302', pwd='gtyhps302'):
    d = Driver()
    d.login(user, pwd)
    activity = d.get_weekly_readers()
    sorted_data = activity_sorting(activity)
    activity['Practice Recording'] = activity['Practice Recording'].apply(lambda x: '0' if x == '-' else x)

    # Define the time duration of 2 hours and 20 minutes
    duration = pd.Timedelta(hours=2, minutes=20)
    weekly_data = activity.iloc[:,[0]].copy()
    weekly_data.loc[:,'Date'] = pd.to_datetime('today').normalize()


    # Create 'Time Completed' column based on condition
    weekly_data.loc[:,'Time Completed'] = activity['Login Time'].apply(lambda x: 'Yes' if x >= duration else 'No')

    # Create 'Recording Completed' column based on condition
    weekly_data.loc[:,'Recording Completed'] = activity['Practice Recording'].apply(lambda x: 'Yes' if int(x) >= 3 else 'No')

    weekly_data.loc[:,'Time Debt'] = activity['Login Time'].apply(lambda x: duration - x if x < duration else 'None')

    # weekly_data['Time Completed'] = activity['Login Time'].apply(lambda x: 'Yes' if x >= duration else 'No')
    # weekly_data['Recordings Completed'] = activity['Practice Recording'].apply(lambda x: 'Yes' if int(x) >= 3  else 'No')

    print("This week's Good RAZ readers are...")

    # print(goods['Students (27)'].to_string())
    tops = sorted_data.tail(5)
    goods = sorted_data[sorted_data['Login Time'] >= duration]
    print(goods.iloc[:, [0]].to_string(index=False),'\n\n\n')

    print("This week's TOP RAZ readers are!")
    print(tops.iloc[:, [0]].to_string(index=False),'\n\n')
    # print(tops['Students (27)'].to_string())

    current_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(current_dir)
    file_path = os.path.join(parent_dir, f'{datetime.date.today()}_{user}_Weekly_Report.csv')
    weekly_data.to_csv(file_path, index=False)
    print(f'{datetime.date.today()}_{user}_Weekly_Report.csv Created')

def program():
    """
    Main function
    :return None:
    """

    print('-' * 10, "Welcome to Mr. Everett's RAZ data toolkit", '-' * 10)
    username = input("Input username:\n")
    password = input("Input password:\n")
    print("'C' for CLASS skills accuracy data over the last 30 days,\n"
          # "'I' for individual student data,\n",
          "or 'W' for WEEKLY time and 3 recording check\n",
          # " *"*8, "WARNING", " *"*8, "\n",
          # "'I' for individual student data takes about 1 minute per student selected\n",
          # "The process is also prone to stalling if internet connection is unstable.",
          # "It will export a PDF with reports for each student for the selected students"
          )
    user_input1 = input()

    try:
        if user_input1.upper() == 'C':
            class_data(username, password) if username.strip() else class_data()
            # os.system('osascript -e \'tell application "Terminal" to activate\'')
        elif user_input1.upper() == 'I':
            student_data(username, password) if username.strip() else student_data()
            # os.system('osascript -e \'tell application "Terminal" to activate\'')
        elif user_input1.upper() == 'W':
            weekly_readers(username, password) if username.strip() else weekly_readers()
            # os.system('osascript -e \'tell application "Terminal" to activate\'')
    except ValueError as ve:
        print(f"Value Error: {ve}")
    return "Program Complete"


title_all = ['Are You a Quitter?', 'Barb and Her Car', 'The Bee and the Flea', 'Big Bad Bat', 'Bob and Nell', 'Boxes and Foxes',
        'Bub and the Nut', 'Chase and Chet', 'Chris the Chef', 'Click, Cluck, and Quack', 'The Clown Who Lost Her Smile',
 'The Club', 'Dan the Tan Man', 'Did It Fit?', 'Don and Dots', 'The Fat Cat', 'A Fat Hat', 'Fixed', 'Fran Goes to the Prom',
 'Fun at Branch Ranch', 'Get the Gag', 'Get the Pets', 'The Girl Who Twirled', 'Goats and Crows', 'Grand Slam Tennis', 'Have You Ever Seen a Fox?',
 'Hot at the Dam', 'I Can Hop', 'Jake and Gail Go to Spain', 'Jazz with Jill', 'Jig, Jag, and Jog', 'Jill and Bill', 'JoJo Gets a Yo-Yo',
 'Kate and Jake', 'The Kind Knight', 'The King Lost His Ring', 'Kit and Kim Are Kin', 'Lil, Sal, and Bill', 'The Mare and the Hare',
 'Midge Gets a Pet', 'The Mutt and the Bug', 'My Pug Has Fun', 'Nan and Pap', 'A Nap and a Map', "Nat's Cat", 'The Nice Mice',
 'No More Sad Tunes', 'Phone Photos', 'Planet Blip Blop', 'The Ramp', 'Red Hen and Rod Rat', 'The Robin', 'Rose the Mole', 'Sam and the Sap',
 'The Shell Shop', 'The Show', 'Skating and Hopping', 'The Storm', 'Swiss Fun Run', 'A Tap and a Pat', 'Ten Pets', 'Thad and Thelma',
 'The Pin with a Tin Fin', 'The Tot and the Pot', 'A Toy for Roy', 'Trip to the City', 'Vin and Val', 'Whisker Bill', 'Win a Wig', 'Wrap It Up',
 'Yum, Yum Yams', 'The Zim-Zam Man']


if __name__ == '__main__':
    program()
    input("Press any key to exit")
    exit()